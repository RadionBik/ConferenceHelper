import string
import openpyxl
import wget
import os
import argparse 


def strip_non_alhpabet(text):
    text = text.replace('.', ' ')
    exclude = set(string.punctuation)
    exclude = exclude | set('td')
    return list(map(lambda x: ''.join(ch for ch in x if ch not in exclude), [text]))[0]

def get_folder_path_from_cell(ws_cell, index):
    ID = ws_cell(row=index, column=1).value
    first_author_surname = ws_cell(row=index, column=2).value.split(' ')[0]
    conference = ws_cell(row=index, column=9).value
    section = strip_non_alhpabet(ws_cell(row=index, column=10).value)
    return f'{conference}/{section}/{ID} {first_author_surname}'.strip() 


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-f", "--file",
        help="xlsx file",
        required=True
        )

    parser.add_argument(
        "-s", "--startrow",
        type=int, 
        help="starting row index",
        required=True
        )

    parser.add_argument(
        "-e", "--endrow",
        type=int, 
        help="ending row index",
        required=True
        )

    args = parser.parse_args()

    START_ROW = args.startrow
    END_ROW = args.endrow
    if START_ROW < 1:
        raise argparse.ArgumentTypeError('Начальный индекс не может быть меньше 1')
    if START_ROW > END_ROW:
        raise argparse.ArgumentTypeError('Конечный индекс не может быть меньше начального')

    wb = openpyxl.load_workbook(args.file)
    ws = wb['Sheet1']
    seen_articles = set()
    # start from the 2nd row
    for index in range(START_ROW, END_ROW+1):
        # consider only first author's rows
        if not ws.cell(row=index, column=1).value:
            continue

        # skip duplicates
        article_name = ws.cell(row=index, column=11).value
        if article_name in seen_articles:
            continue
        seen_articles.add(article_name)
         
         
        article_folder = get_folder_path_from_cell(ws.cell, index)        
        if os.path.exists(article_folder):
            continue

        print(f'\nСкачиваем файлы в папку: {article_folder}:')
        os.makedirs(article_folder)
            
        for link_column in [14, 15, 16, 18, 19]:
            try:
                link = ws.cell(row=index, column=link_column).hyperlink.target
            except AttributeError:
                continue

            wget.download(link, article_folder)

        index+=1

if __name__ == '__main__':
    main()

